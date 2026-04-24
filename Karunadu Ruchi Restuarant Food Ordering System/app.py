import os
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, session
from openpyxl import load_workbook, Workbook

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Change for production

EXCEL_FILE = 'data/orders.xlsx'

# Ensure data directory exists
os.makedirs('data', exist_ok=True)

def init_excel():
    """Create Excel file with two sheets if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        # Users sheet
        ws_users = wb.active
        ws_users.title = 'Users'
        ws_users.append(['Name', 'Phone', 'Login Time'])
        # Orders sheet
        ws_orders = wb.create_sheet('Orders')
        ws_orders.append(['Name', 'Phone', 'Order Time', 'Items', 'Total (₹)'])
        wb.save(EXCEL_FILE)

def save_login(name, phone):
    """Append login info to Users sheet."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb['Users']
    ws.append([name, phone, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    wb.save(EXCEL_FILE)

def save_order(name, phone, items, total):
    """Append order details to Orders sheet."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb['Orders']
    # Convert items list to a string for Excel storage
    items_str = ', '.join([f"{it['name']} ({it['qty']} x ₹{it['price']})" for it in items])
    ws.append([name, phone, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), items_str, total])
    wb.save(EXCEL_FILE)

# Extended menu with image URLs (replace with your own images)
MENU = [
    {'id': 1, 'name': 'Shahi Paneer', 'price': 150,'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSpJ6N_2tTLrZiA7XDv2d-HtbSf7Ho8jqo5tw&s'},
    {'id': 2, 'name': 'palak paneer', 'price': 100, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcTFXg7a7-y1VgdTYdThObeQEIhln08gIKWDFA&s'},
    {'id': 3, 'name': 'Dal Tadka', 'price': 100, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQYybic8KcylV4XUamukL-qhaj0bbbyghSKbA&s'},
    {'id': 4, 'name': 'Veg Biryani', 'price': 90, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRPbQMgD79WJwULTkrE5C-MsHcUMyKozMq-Kw&s'},
    {'id': 5, 'name': 'Paneer Butter Masala', 'price': 101, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcTZ1kcoiXbqzazXFd2VrQPaheS9YU7Bss8eHw&s'},
    {'id': 6, 'name': 'Rajma Chawal', 'price': 80, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcShVmohcGGjRtkSCn2ws28RKeAzFKBGgXnrcQ&s'},
    {'id': 7, 'name': 'Chole Masala', 'price': 65, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcTh0wY-WVTJ0X7box1spdWZHT-X8IbQ5gntdA&s'},
    {'id': 8, 'name': 'Aloo Gobi', 'price': 70, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQr0amxAjBENxG-VPzWY2WQ1d5NIVyt2ZCulA&s'},
    {'id': 9, 'name': 'Malai Kofta','price':90,'image': 'https://thumbs.dreamstime.com/b/malai-kofta-indian-vegetarian-meatballs-curry-traditionally-served-hot-flatbread-tandoori-rumali-roti-naan-traditional-158606665.jpg'},
    {'id': 10, 'name': 'Veg Pulao','price':75 ,'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRXh8JwS9wEk8r2un31eZJtUqoEIveNLTHibg&s'},
    {'id': 11, 'name': 'Masala Dosa','price':70,'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSKGOuuXW3ot708IAtRqaCEBntRuBHHiLKM8Q&s'},
    {'id': 12, 'name': 'Idli Sambar','price':40,'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRthFRj3vH0p3p6NRcMkBldEezsaEPGATN6kw&s'},
    {'id': 13, 'name': 'Poha', 'price': 50,'image': 'https://www.mrishtanna.com/wp-content/uploads/2018/04/poha-indian-breakfast-recipe.jpg'},
    {'id':14, 'name': 'Upma', 'price': 30,'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSLc3M7CabPKV0G5OH7QaOmQ8jQARiMyw5oRg&s'},
    {'id':15, 'name': 'Aloo Paratha', 'price': 50, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcR8WJXKEF9A74OvfeTRGSRf4o78xFESSOnfOA&s'},
    {'id': 16, 'name': 'Vada Sambar', 'price': 25, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcTuQOO_r5OW0dKLIjMkz6NvZ5GIWpUl4qLiIw&s'},
    {'id': 17, 'name': 'Pongal', 'price': 30, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcS3GVuGl05tIZXNdh1ib-LOOJa-m_Gw80Y_0Q&s'},
    {'id': 18, 'name': 'Veg Sandwich', 'price':50, 'image': 'https://thumbs.dreamstime.com/b/veg-grilled-sandwich-served-ketchup-isolated-over-rustic-wooden-background-selective-focus-224440470.jpg'},
    {'id': 19, 'name': 'Puri Bhaji', 'price': 50, 'image':'https://img.freepik.com/free-photo/delicious-indian-culinary-experience_23-2151998597.jpg?semt=ais_hybrid&w=740&q=80'},
    {'id': 20, 'name': 'Uttappam', 'price':40, 'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcR4sZ_qSL1JOyBnidVQ6zgiKS9ZNIiJiWTOlQ&s'},
    {'id': 21, 'name': 'Chicken Biriyani', 'price':99, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRbaSN9yaFJNMirjoxpvIWAsDhOvQUD9fgA5g&s'},
    {'id': 22, 'name': 'Butter Chicken', 'price':120, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcTOmHzb0vLeLbuc7NiP3w5y2J30-NkBRPHrKw&s'},
    {'id': 23, 'name': 'Chicken Curry', 'price':110, 'image': 'https://static.vecteezy.com/system/resources/thumbnails/038/972/486/small/ai-generated-chicken-korma-isolated-on-white-indian-cuisine-meat-curry-dish-with-coconut-milk-masala-asian-food-photo.jpg'},
    {'id': 24, 'name': 'Fish Curry', 'price':100, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRkmOKogYC-nhKy3wcnspsgz1feQn3WZ8IZ1g&s'},
    {'id':25, 'name': 'Egg Curry', 'price':75, 'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQhOJR6mhkr8a0bzhUo1oLis5PTxqFWmnYfOQ&s'},
    {'id':26, 'name': 'Chicken Tikka', 'price':120, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQb2TiZw63iyY7S7rzHPTBDzg5r9Rnd0lvHYA&s'},
    {'id':27, 'name': 'Prawn Masala', 'price':110, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQ4zth2I6sTQTunNTdYgTy9VphVGyDhtJo27w&s'},
    {'id':28, 'name': 'Tandoori Chicken', 'price':200, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRk79LP7Jt4-28FBnViEY3xH1VyeDi2AG379Q&s'},
    {'id':29, 'name': 'Chicken Fried Rice', 'price':100, 'image': 'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSdYscNsOtcOqTmjr0kQES62GqlrNLyk3D7Lg&s'},
    {'id':30, 'name': 'Paneer Tikka', 'price':120, 'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQnuTuasyQlQv4teSGHeczcS8aB7Yko59CzvQ&s'},
    {'id':31, 'name': 'Veg Manchurian', 'price':90,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcTb2lR88Psee9AwIk2YuVuYSs8lTdvBkHhEpw&s'},
    {'id':32, 'name': 'Gobi Manchurian', 'price':70,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSHRESiYmwBMp2OS85ffLQN4SYrRJIcyYqJag&s'},
    {'id':33, 'name': 'Chilli Chicken', 'price':110,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcThN2etbfWCvYoN-jsMLc0522fdTF3b7UVuiw&s'},
    {'id':34, 'name': 'Spring Rolls', 'price':60,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRz4Ji139n10z90LIx0g_IzCGtcB99QhHT2Ew&s'},
    {'id':35, 'name': 'Chicken Lollipop', 'price':120,'image':'https://t3.ftcdn.net/jpg/04/07/37/94/360_F_407379443_KIdJDRuKDLOSLNT93VN3PfFh8hzGEO9P.jpg'},
    {'id':36, 'name': 'Crispy Corn', 'price':70,'image':'https://rakskitchen.net/wp-content/uploads/2022/01/crisp-corn-500x375.jpg'},
    {'id':37, 'name': 'Honey Chilli Potato', 'price':70,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQvIDh8E8Pi8EWxu5Na_T3Nq1lhSROQQI0B4A&s'},
    {'id':38, 'name': 'Fish Fingers', 'price':100,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcT3PA-TLFzCeiQJPtof5POFYdUFX2Pnz4eqiQ&s'},
    {'id':39, 'name': 'Chicken Kabab', 'price':120,'image':'https://i.ytimg.com/vi/bkLqg1fH-AY/hq720.jpg?sqp=-oaymwEhCK4FEIIDSFryq4qpAxMIARUAAAAAGAElAADIQj0AgKJD&rs=AOn4CLAEOl2qH8_aWVwIuaooDz8oKXbAlQ'},
    {'id':40, 'name': 'Samosa', 'price':20,'image':'https://t4.ftcdn.net/jpg/04/66/42/25/360_F_466422564_LICnIvfjfGhieSKG4gxU35LirfjrxbOB.jpg'},
    {'id':41, 'name': 'Vada Pav', 'price':30,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQVptCP0V3kIacit1Lsv4099Ftjpijb8DKuNw&s'},
    {'id':42, 'name': 'Pani Puri', 'price':40, 'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSDoyCIacfjPpaA08SOLaoWC7Amq1WoNXjYNQ&s'},
    {'id':43, 'name': 'Bhel Puri', 'price':60, 'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSgZQqvJOFsxI5gqPuXg0cohm3qjp98VBJG3Q&s'},
    {'id':44, 'name': 'Pakora', 'price':60, 'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQog0Y5anS6M-sIUGK3CTnqLrixrYirX3VBAQ&s'},
    {'id':45, 'name': 'Bread Pakora', 'price':40,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcR1Z_U0zcs5L3p58gDsxJlUOOsNRBE4fCSRLQ&s'},
    {'id':46, 'name': 'Cutlet', 'price':50,'image':'https://i0.wp.com/veenapatwardhan.com/pat-a-cake/wp-content/uploads/2023/02/Vegetable-Cutlets-New-Way.jpg?fit=760%2C400&ssl=1'},
    {'id':47, 'name': 'Maggi', 'price':55, 'image':'https://t4.ftcdn.net/jpg/03/67/31/89/360_F_367318954_4ttRBcmaa22q30OsJRqflxgCOd9dB9t6.jpg'},
    {'id':48, 'name': 'Corn Chatt', 'price':45,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcToPOhfOBNsgpzOHczZN4FmGyeo3jGMpafJ5g&s'},
    {'id':49, 'name': 'Kachori', 'price':30,'image':'https://thumbs.dreamstime.com/b/kachori-chutney-indian-snack-60979604.jpg'},
    {'id':50, 'name': 'Mango Juice','price':55,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSpnxG_55J8vAfv5BZyQ3zjI6V8EMAOa6RluQ&s'},
    {'id':51, 'name': 'Orange Juice', 'price':45,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRIfh0-g0TYH4CLKXkpjxsEaPH-DppDxmewJQ&s'},
    {'id':52, 'name': 'Watermelon Juice', 'price':35,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSFsCfF1PSvzsUsoMxTahyOyYg4p0fkIMDKOQ&s'},
    {'id':53, 'name': 'Pineapple Juice', 'price':40, 'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcTKf8i8C8JzfcEm8Oa2-DU9yF71UTzZPGDoig&s'},
    {'id':54, 'name': 'Mosamabi Juice', 'price':45, 'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSzFAoi4U9s9d__0gli4AwnDKtdiFcDOvLO8w&s'},
    {'id':55, 'name': 'Pomegranate Juice', 'price':50,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcT-fsaLih5LZl4yzUwwPjEC7sj24BgHiQMkLA&s'},
    {'id':56, 'name': 'Apple Juice', 'price':60, 'image':'https://t3.ftcdn.net/jpg/03/07/37/78/360_F_307377873_MVxS5nP1MnaoKYCfwqCFrXWunwGhjG8v.jpg'},
    {'id':57, 'name': 'Carrot Juice', 'price':56,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRMAR55uBuJlNtAZeDVuUcXj9vG3mdIXZ9RsQ&s'},
    {'id':58, 'name': 'Mixed Fruit Juice', 'price':80, 'image':'https://5.imimg.com/data5/SELLER/Default/2022/8/YS/AN/QA/124465290/mixed-fruit-juice-500x500.jpg'},
    {'id':59, 'name': 'Sugarcane Juice', 'price':30, 'image':'https://upload.wikimedia.org/wikipedia/commons/thumb/f/f9/Glass_of_sugarcane_juice.jpg/960px-Glass_of_sugarcane_juice.jpg'},
    {'id':60, 'name': 'Vanilla Icecream','price':60, 'image':'https://media.istockphoto.com/id/1326143969/photo/bowl-with-vanilla-ice-cream-balls.jpg?s=612x612&w=0&k=20&c=WxEriNEK7yW7F4AWImLQRrpco-R_bdDYEQoyhigu9fc='},
    {'id':61, 'name': 'Chocolate Icecream','price':70,'image':'https://media.istockphoto.com/id/1491655936/photo/chocolate-ice-cream-scoop.jpg?s=612x612&w=0&k=20&c=m_NDyHXn048dbKDuAfcnRwP54AM74xQMCWsUlOL-JJI='},
    {'id':62,'name':'Strawberry Icecream','price':70,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRT7iio8v21SWHMHBfSiaed0Ti7W1L9ZH8vHQ&s'},
    {'id':63,'name':'Butterscotch Icecream','price':80,'image':'https://t4.ftcdn.net/jpg/07/04/72/81/360_F_704728190_tqYFHbBEPDl8KlFWEbcIaN5xzzP4Ho3G.jpg'},
    {'id':64,'name':'Black Currant Icecream','price':80,'image':'https://5.imimg.com/data5/SELLER/Default/2021/3/EV/LS/IE/60135453/custard-apple-ice-cream.jpg'},
    {'id':65,'name':'Mango Icecream','price':57,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRJtrV_HjwosZzcfJ4JJue_5l548ZSfyr2LWA&s'},
    {'id':66,'name':'Kulfi Icecream','price':40,'image':'https://5.imimg.com/data5/CC/AK/MY-69718924/kulfi-ice-cream-500x500.jpg'},
    {'id':67,'name':'Choco chip Icecream','price':50,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRBQUI6j93HF2icF0XPFxwxFkZcLdAy1v0n5A&s'},
    {'id':68,'name':'Pista Icecream','price':67,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcR7g6aWv_SqHKIlu5VsGDyUstCwpPz5a2kyuA&s'},
    {'id':69,'name':'Cassata Icecream','price':55,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQiOlmaVmx9PwdsZLNn4r7UOXHjPEiCk4wAPw&s'},
    {'id':70,'name':'Cup Icecream','price':10,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRCwSY-JO46dWhdTLSMS5pwuijNMGTIfepCmw&s'},
    {'id':71,'name':'Masala Tea','price':20,'image':'https://t4.ftcdn.net/jpg/02/97/69/23/360_F_297692313_RO2FGXo1qTq6jH4sOOVMCWYGnAuYLR3d.jpg'},
    {'id':72,'name':'Green Tea','price':50,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRq_ksokpjOm1RHqtR1PFu2Mm22dSZkBHu3Lg&s'},
    {'id':73,'name':'Ginger Tea','price':30,'image':'https://t4.ftcdn.net/jpg/03/57/34/23/360_F_357342355_FIhhTeOlzO2xObSK20RTIOOywpXikdPp.jpg'},
    {'id':74,'name':'Lemon Tea','price':25,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRgsnhS7PL8m2pkdWdHbu-G0nJbWjzwt2Vf6A&s'},
    {'id':75,'name':'Black Tea','price':30,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcTc-RHtxmdjaqTjO-8m989Z8kxHaExSrMfjQA&s'},
    {'id':76,'name':'Elaichi Tea','price':20,'image':'https://shiatea.com/cdn/shop/files/cardamomchai9.jpg?v=1706084387&width=1445'},
    {'id':77,'name':'Tulsi Tea','price':25,'image':'https://4.imimg.com/data4/YH/RR/MY-26339309/tulsi-tea-500x500.jpg'},
    {'id':78,'name':'Kashmiri Kahwa Tea','price':39,'image':'https://www.cookclickndevour.com/wp-content/uploads/2021/01/Kashmiri-kahwa-tea-recipe-2.jpg'},
    {'id':79,'name':'Iced Tea','price':45,'image':'https://t3.ftcdn.net/jpg/08/48/92/02/360_F_848920225_ShRI77tjoshBJ1y1wFHfbbZk1xkI0fhA.jpg'},
    {'id':80,'name':'Milk Tea','price':15,'image':'https://t3.ftcdn.net/jpg/01/06/14/76/360_F_106147605_xmXfzxpraUtLQZkwTYWhjIJc0dkBkN8Y.jpg'},
    {'id':81,'name':'Filter Coffee','price':40,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQi7dRMX1XOk2tcPEkQ6iJvvAJ-J6-M0Jjjfg&s'},
    {'id':82,'name':'Espresso Coffee','price':45,'image':'https://img.freepik.com/free-photo/closeup-classic-fresh-espresso-served-dark-surface_1220-5375.jpg?semt=ais_hybrid&w=740&q=80'},
    {'id':83,'name':'Cappuccino Coffee','price':50,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcR3JCS9lrOqbrfD1HQr1fwS40OYMIy-XOOF7w&s'},
    {'id':84,'name':'Latte Coffee','price':55,'image':'https://img.freepik.com/free-photo/latte-coffee_74190-7821.jpg?semt=ais_hybrid&w=740&q=80'},
    {'id':85,'name':'Cold Coffee','price':60,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcS82M8mUlk1UgN1Dcv9Njnq9-iu-ZcMyAM8kA&s'},
    {'id':86,'name':'Mocha Coffee','price':65,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRefU6BIUDyTAORf1o8WrByF7poO_Wnw_c1Cw&s'},
    {'id':87,'name':'Americano Coffee','price':50,'image':'https://img.freepik.com/premium-photo/classic-americano-coffee-with-milk-gray-background_112304-1304.jpg'},
    {'id':88,'name':'Irish Coffee','price':70,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcT7rqBmxNCnY9OC_pIpFu59kA8pRF7TzdEnEw&s'},
    {'id':89,'name':'Caramel Coffee','price':45,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcR2L6wIhbXLeDplhwW_UWpBURg09R1pCRGrcQ&s'},
    {'id':90,'name':'Iced Coffee','price':69,'image':'https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSKIEXCoefXcGXZONVk48FxHTD2qjQqYbfwWA&s'}
]

@app.route('/')
def index():
    return render_template('index.html', now_year=datetime.now().year)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        session['name'] = name
        session['phone'] = phone
        init_excel()
        save_login(name, phone)
        session['cart'] = []
        return redirect(url_for('menu'))
    return render_template('login.html')

@app.route('/menu')
def menu():
    if 'name' not in session:
        return redirect(url_for('login'))
    return render_template('menu.html', menu=MENU)

@app.route('/add_to_cart', methods=['POST'])
def add_to_cart():
    data = request.get_json()
    item_id = int(data['item_id'])
    quantity = int(data['quantity'])
    item = next((i for i in MENU if i['id'] == item_id), None)
    if item and quantity > 0:
        cart = session.get('cart', [])
        for cart_item in cart:
            if cart_item['id'] == item_id:
                cart_item['qty'] += quantity
                break
        else:
            cart.append({
                'id': item['id'],
                'name': item['name'],
                'price': item['price'],
                'qty': quantity
            })
        session['cart'] = cart
        return {'success': True, 'cart': cart}
    return {'success': False}, 400

@app.route('/cart')
def cart():
    if 'name' not in session:
        return redirect(url_for('login'))
    cart_items = session.get('cart', [])
    total = sum(item['price'] * item['qty'] for item in cart_items)
    return render_template('cart.html', cart=cart_items, total=total)

@app.route('/place_order', methods=['POST'])
def place_order():
    if 'name' not in session:
        return redirect(url_for('login'))
    cart_items = session.get('cart', [])
    if not cart_items:
        return redirect(url_for('menu'))
    total = sum(item['price'] * item['qty'] for item in cart_items)
    name = session['name']
    phone = session['phone']
    init_excel()
    save_order(name, phone, cart_items, total)
    session.pop('cart', None)
    session['last_order'] = {
        'items': cart_items,
        'total': total
    }
    return redirect(url_for('bill'))

@app.route('/bill')
def bill():
    if 'name' not in session or 'last_order' not in session:
        return redirect(url_for('login'))
    name = session['name']
    phone = session['phone']
    order = session['last_order']
    return render_template('bill.html', name=name, phone=phone, items=order['items'], total=order['total'])

if __name__ == '__main__':
    app.run(debug=True)